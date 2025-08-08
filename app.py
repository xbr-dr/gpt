# app.py - CampusGPT (Batch 1: Header & Configuration)
import os
import sys
import time
import json
import pickle
import shutil
import tempfile
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Any

# Core libs for app
import streamlit as st

# Data processing & ML helpers (used in later batches)
import numpy as np
import pandas as pd
import faiss
import nltk
import re

# File reading
from PyPDF2 import PdfReader
from openpyxl import load_workbook

# Optional: sentence-transformers, geopy, folium, groq will be imported in Batch 2
# to allow faster incremental testing if you want to run partial batches.



# ------------------- Batch 3: File processing, extraction, chunking, indexing -------------------
from typing import Iterable
# ------------------- Batch 4: Intent classification, retrieval, re-ranking, LLM prompt -------------------
import datetime
import hashlib
import math
# ------------------- Batch 5: Map UI, folium integration, map mode toggle & focus -------------------
import folium
from streamlit_folium import st_folium
from geopy.distance import geodesic
# ------------------- Batch 6: UI pages (User & Admin) and main app wiring -------------------
import uuid
import glob
# ------------------- Batch 7: UI polish, structured prompts, admin utilities -------------------
import html
import time as _time
import streamlit as st
GROQ_API_KEY = st.secrets["GROQ_API_KEY"]

# ------------------- Batch 2: Models, NLTK, LangDetect, Location loading -------------------
import threading
from langdetect import detect, DetectorFactory, LangDetectException

# Try to import optional fast fuzzy library (rapidfuzz). Fallback to difflib if not available.
try:
    from rapidfuzz import fuzz, process
    RAPIDFUZZ_AVAILABLE = True
except Exception:
    from difflib import get_close_matches
    RAPIDFUZZ_AVAILABLE = False

# ML model & Groq imports (may be heavy)
try:
    from sentence_transformers import SentenceTransformer
except Exception as e:
    SentenceTransformer = None
    _log_debug(f"SentenceTransformer import failed: {e}")

try:
    from groq import Groq
except Exception as e:
    Groq = None
    _log_debug(f"Groq import failed: {e}")
# ------------------- Configuration -------------------
# Directories
BASE_DIR = Path.cwd()
DATA_DIR = BASE_DIR / "data"
DOCUMENTS_DIR = DATA_DIR / "documents"
LOCATIONS_DIR = DATA_DIR / "locations"  # optional for storing CSV and backups
STORAGE_DIR = BASE_DIR / "storage"

# Ensure directories exist
for p in (DOCUMENTS_DIR, LOCATIONS_DIR, STORAGE_DIR):
    p.mkdir(parents=True, exist_ok=True)

# Filenames / Paths (persisted artifacts)
FAISS_INDEX_PATH = STORAGE_DIR / "faiss_index.faiss"
CORPUS_PATH = STORAGE_DIR / "corpus.pkl"
LOCATION_DATA_PATH = STORAGE_DIR / "locations.pkl"
CHAT_HISTORY_DIR = STORAGE_DIR / "chat_sessions"  # JSON files per session

CHAT_HISTORY_DIR.mkdir(parents=True, exist_ok=True)

# Admin credentials
# - Production note: store secrets in Streamlit secrets or environment variables.
DEFAULT_ADMIN_PASSWORD = "1234"  # fallback; override in streamlit secrets
ADMIN_PASSWORD = os.environ.get("CAMPUSGPT_ADMIN_PASSWORD") or st.secrets.get("admin_password", DEFAULT_ADMIN_PASSWORD)

# Model / service config placeholders (will be initialized in Batch 2)
SENTENCE_TRANSFORMER_MODEL = "distiluse-base-multilingual-cased-v1"
GROQ_MODEL_NAME = "llama3-8b-8192"

# Application settings
MAX_RETRIEVE = 6              # number of RAG chunks to retrieve by default
MIN_SENTENCE_LENGTH = 25      # minimum chars to keep a sentence during chunking
MAX_CHAT_HISTORY_ITEMS = 200  # limit stored session messages to avoid bloat

# Behavior configuration: strict domain-only policy
STRICT_DOMAIN_ONLY = True

# UI / UX constants
APP_TITLE = "CampusGPT — Campus Assistant"
APP_ICON = "🏫"
PAGE_LAYOUT = "wide"

# Streamlit page config (safe to run now)
st.set_page_config(page_title=APP_TITLE, page_icon=APP_ICON, layout=PAGE_LAYOUT, initial_sidebar_state="collapsed")

# Lightweight helper for debug logging on Streamlit (toggle with env)
DEBUG = os.environ.get("CAMPUSGPT_DEBUG", "0") == "1"

def _log_debug(msg: str):
    if DEBUG:
        st.text(f"[DEBUG] {msg}")

_log_debug("Batch 1 loaded: Header & Configuration")



# Geopy & mapping libs will be imported in later batches when map rendering is added.

# ------------------- NLTK setup -------------------
# Make langdetect deterministic
DetectorFactory.seed = 0

@st.cache_resource
def load_nltk_and_tokenizer() -> bool:
    """
    Ensure NLTK punkt tokenizer is available. Returns True if ready.
    Cached as a resource to avoid re-downloads.
    """
    try:
        nltk.data.find("tokenizers/punkt")
        _log_debug("NLTK punkt already available")
        return True
    except LookupError:
        try:
            nltk.download("punkt", quiet=True)
            nltk.data.find("tokenizers/punkt")
            _log_debug("NLTK punkt downloaded successfully")
            return True
        except Exception as e:
            st.error(f"❌ Failed to download NLTK data: {e}")
            return False

NLTK_READY = load_nltk_and_tokenizer()

# ------------------- Model & Groq loader -------------------
@st.cache_resource
def load_models_and_groq(model_name: str = SENTENCE_TRANSFORMER_MODEL, groq_model_name: str = GROQ_MODEL_NAME):
    """
    Loads the sentence transformer model and initializes Groq client if API key present.
    Returns (embed_model, groq_client) where either may be None on failure.
    """
    embed_model = None
    groq_client = None

    # Load sentence-transformer embedding model
    if SentenceTransformer is None:
        st.warning("SentenceTransformer library not found. Install 'sentence-transformers' to enable semantic search.")
    else:
        try:
            embed_model = SentenceTransformer(model_name)
            _log_debug(f"Loaded embedding model: {model_name}")
        except Exception as e:
            st.error(f"❌ Failed to load embedding model '{model_name}': {e}")
            embed_model = None

    # Initialize Groq client if possible
    groq_api_key = os.environ.get("GROQ_API_KEY") or st.secrets.get("GROQ_API_KEY")
    if not groq_api_key:
        st.warning("Groq API key not set. LLM responses will be disabled until a valid key is provided.")
    else:
        if Groq is None:
            st.error("Groq client library not found. Install 'groq' to enable LLM features.")
        else:
            try:
                groq_client = Groq(api_key=groq_api_key)
                _log_debug("Groq client initialized")
            except Exception as e:
                st.error(f"❌ Failed to initialize Groq client: {e}")
                groq_client = None

    return embed_model, groq_client

# Load at startup (cached)
EMBED_MODEL, GROQ_CLIENT = load_models_and_groq()

# ------------------- Location CSV loader -------------------
def _generate_acronym(name: str) -> str:
    """Generate a simple acronym for a multi-word name like 'Information Technology' -> 'IT'."""
    parts = re.findall(r"[A-Za-z]+", name)
    if not parts:
        return ""
    if len(parts) == 1:
        # Take first 2 letters for single-word names up to 3 letters
        return parts[0][:2].upper()
    return "".join(p[0].upper() for p in parts if p)

def _normalize_aliases(raw_aliases: Optional[str], primary_name: str) -> List[str]:
    """
    Parse aliases string (pipe-separated) and add generated acronyms and normalized forms.
    Returns a list of lowercased aliases without duplicates.
    """
    aliases = set()
    if raw_aliases and isinstance(raw_aliases, str):
        for a in re.split(r"\||,|;", raw_aliases):
            a = a.strip()
            if a:
                aliases.add(a.lower())

    # Add normalized primary name and simple variants
    aliases.add(primary_name.lower())
    aliases.add(primary_name.replace(",", "").lower())

    # Add acronym
    acronym = _generate_acronym(primary_name)
    if acronym:
        aliases.add(acronym.lower())
        # also "dept of X" style
        aliases.add(f"{acronym.lower()} department")
        aliases.add(f"department of {acronym.lower()}")

    # Add "department of" variants for multi-word names
    aliases.add(f"department of {primary_name.lower()}")
    aliases.add(f"{primary_name.lower()} department")
    aliases.add(f"dept of {primary_name.lower()}")

    # filter empties and return
    return sorted({a for a in aliases if a})

@st.cache_resource
def load_locations_from_csv(csv_path: Optional[Path] = None) -> Dict[str, Dict[str, Any]]:
    """
    Loads locations from a CSV in LOCATIONS_DIR or given path.
    Expected columns: name, latitude/lat, longitude/lon, description (aliases optional).
    Returns a dict keyed by normalized primary name -> location dict with 'aliases' list.
    """
    path_to_try = csv_path or (LOCATIONS_DIR / "locations.csv")
    locations_map: Dict[str, Dict[str, Any]] = {}

    if not path_to_try.exists():
        _log_debug(f"No locations CSV found at {path_to_try}")
        return {}

    try:
        df = pd.read_csv(path_to_try)
        # Normalize column names
        df.columns = [str(c).strip().lower() for c in df.columns]

        # Identify column names flexibly
        name_col = next((c for c in df.columns if c in ('name', 'place', 'location')), None)
        lat_col = next((c for c in df.columns if c in ('latitude', 'lat', 'y')), None)
        lon_col = next((c for c in df.columns if c in ('longitude', 'lon', 'x')), None)
        aliases_col = next((c for c in df.columns if c in ('aliases', 'alias', 'nicknames')), None)
        desc_col = next((c for c in df.columns if c in ('description', 'desc', 'details')), None)

        if not (name_col and lat_col and lon_col):
            st.warning("Locations CSV missing required columns. Expected at least: name, latitude, longitude.")
            return {}

        for _, row in df.iterrows():
            try:
                primary_name = str(row[name_col]).strip()
                if not primary_name:
                    continue

                lat = float(row[lat_col])
                lon = float(row[lon_col])

                raw_aliases = str(row[aliases_col]) if aliases_col and pd.notna(row.get(aliases_col)) else ""
                desc = str(row[desc_col]) if desc_col and pd.notna(row.get(desc_col)) else ""

                normalized_key = primary_name.strip().lower()
                alias_list = _normalize_aliases(raw_aliases, primary_name)

                locations_map[normalized_key] = {
                    "name": primary_name,
                    "lat": lat,
                    "lon": lon,
                    "desc": desc,
                    "aliases": alias_list,
                    "original_name": primary_name
                }
            except Exception as ex_row:
                _log_debug(f"Skipping invalid location row: {ex_row}")
                continue

        # Persist processed locations for fast load
        with open(LOCATION_DATA_PATH, "wb") as f:
            pickle.dump(locations_map, f)
        _log_debug(f"Loaded {len(locations_map)} locations from CSV")

    except Exception as e:
        st.error(f"❌ Failed to read locations CSV: {e}")
        return {}

    return locations_map

# Try to load any existing saved location map (prefers pickled cached version)
@st.cache_resource
def get_location_map() -> Dict[str, Dict[str, Any]]:
    # If pickle exists, load it for speed; else try csv loader
    if LOCATION_DATA_PATH.exists():
        try:
            with open(LOCATION_DATA_PATH, "rb") as f:
                locs = pickle.load(f)
                _log_debug(f"Loaded {len(locs)} locations from pickle cache")
                return locs
        except Exception as e:
            _log_debug(f"Failed to load locations pickle: {e}")
    # attempt to read CSV
    return load_locations_from_csv()

LOCATION_MAP = get_location_map()

# Fuzzy match helper that uses rapidfuzz if available, else difflib
def fuzzy_match_location(query: str, location_map: Dict[str, Dict[str, Any]], score_cutoff: int = 75) -> Optional[Dict[str, Any]]:
    """
    Try to match query to a location (name or aliases).
    Returns the matched location dict or None if not found.
    score_cutoff is in 0-100 (higher stricter).
    """
    if not query or not location_map:
        return None

    q = query.lower().strip()

    # First, direct substring search on keys and aliases
    for key, loc in location_map.items():
        if key in q:
            return loc
        for alias in loc.get("aliases", []):
            if alias in q:
                return loc

    # Build list of candidate names (primary + aliases)
    candidates = []
    for key, loc in location_map.items():
        candidates.append((key, loc))
        for a in loc.get("aliases", []):
            candidates.append((a, loc))

    # Use rapidfuzz/process.extractOne if available
    try:
        if RAPIDFUZZ_AVAILABLE:
            names = [c[0] for c in candidates]
            best = process.extractOne(q, names, scorer=fuzz.token_sort_ratio)
            if best and best[1] >= score_cutoff:
                matched_name = best[0]
                # find loc for matched_name
                for name, loc in candidates:
                    if name == matched_name:
                        return loc
        else:
            # difflib fallback: try get_close_matches on candidate names
            names = [c[0] for c in candidates]
            matches = get_close_matches(q, names, n=1, cutoff=score_cutoff/100.0)
            if matches:
                matched_name = matches[0]
                for name, loc in candidates:
                    if name == matched_name:
                        return loc
    except Exception as e:
        _log_debug(f"Fuzzy matching failed: {e}")

    return None

_log_debug("Batch 2 loaded: Models (lazy), NLTK ready, Location loader ready")



# Helper: allowed file extensions
ALLOWED_UPLOAD_TYPES = {'.pdf', '.txt', '.csv', '.xlsx', '.xls'}

def _safe_read_text_from_pdf(path: Path) -> str:
    """Extract text from PDF robustly, skip pages with issues."""
    try:
        reader = PdfReader(str(path))
        texts = []
        for page in reader.pages:
            try:
                ptxt = page.extract_text()
                if ptxt:
                    texts.append(ptxt)
            except Exception:
                continue
        return "\n".join(texts)
    except Exception as e:
        _log_debug(f"PDF read failed for {path}: {e}")
        return ""

def _safe_read_text_from_txt(uploaded_file) -> str:
    try:
        raw = uploaded_file.getvalue()
        if isinstance(raw, bytes):
            return raw.decode('utf-8', errors='ignore')
        return str(raw)
    except Exception as e:
        _log_debug(f"TXT read failed: {e}")
        return ""

def _safe_read_text_from_excel(uploaded_file) -> str:
    """
    Convert excel sheets to concatenated string representation.
    Keeps column names and rows for RAG ingestion.
    """
    try:
        df = pd.read_excel(uploaded_file, engine='openpyxl')
        return df.to_string(index=False)
    except Exception as e:
        _log_debug(f"Excel read failed: {e}")
        # try openpyxl lower-level if needed
        try:
            wb = load_workbook(uploaded_file, read_only=True)
            texts = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    rows.append([str(x) if x is not None else "" for x in row])
                texts.append(f"Sheet: {sheet}\n" + "\n".join([", ".join(r) for r in rows]))
            return "\n\n".join(texts)
        except Exception as e2:
            _log_debug(f"Excel fallback failed: {e2}")
            return ""

def _safe_read_text_from_csv(uploaded_file) -> str:
    try:
        # pandas can read a file-like object
        df = pd.read_csv(uploaded_file)
        return df.to_string(index=False)
    except Exception as e:
        _log_debug(f"CSV read failed: {e}")
        try:
            uploaded_file.seek(0)
            content = uploaded_file.getvalue().decode('utf-8', errors='ignore')
            return content
        except Exception:
            return ""

def process_uploaded_files(uploaded_files: Iterable) -> Tuple[List[Dict[str, str]], Dict[str, Dict]]:
    """
    Accepts list of Streamlit UploadedFile objects.
    Returns (file_data_list, locations_from_files)
    - file_data_list: list of {'text':..., 'source': filename}
    - locations_from_files: dict of found location entries from structured files (CSV/excel)
    """
    file_data = []
    locations_found = {}

    for uploaded in uploaded_files:
        fname = uploaded.name
        ext = os.path.splitext(fname)[1].lower()
        if ext not in ALLOWED_UPLOAD_TYPES:
            st.warning(f"Skipping unsupported file type: {fname}")
            continue

        # Save a copy to documents dir for record
        save_path = DOCUMENTS_DIR / fname
        try:
            with open(save_path, "wb") as f:
                f.write(uploaded.getbuffer())
        except Exception as e:
            st.error(f"Failed to save uploaded file {fname}: {e}")
            continue

        text = ""
        try:
            if ext == '.pdf':
                text = _safe_read_text_from_pdf(save_path)

            elif ext == '.txt':
                text = _safe_read_text_from_txt(uploaded)

            elif ext in ('.xlsx', '.xls'):
                text = _safe_read_text_from_excel(uploaded)
                # Try to extract structured location rows if present
                try:
                    df = pd.read_excel(uploaded, engine='openpyxl' if ext == '.xlsx' else None)
                    extracted = _extract_locations_from_df(df)
                    locations_found.update(extracted)
                except Exception as e:
                    _log_debug(f"No structured locations extracted from {fname}: {e}")

            elif ext == '.csv':
                text = _safe_read_text_from_csv(uploaded)
                try:
                    df = pd.read_csv(uploaded)
                    extracted = _extract_locations_from_df(df)
                    locations_found.update(extracted)
                except Exception as e:
                    _log_debug(f"No structured locations extracted from CSV {fname}: {e}")
        except Exception as e:
            st.error(f"Error processing {fname}: {e}")
            continue

        if text and len(text.strip()) > 20:
            file_data.append({"text": text, "source": fname})

    return file_data, locations_found

def _extract_locations_from_df(df: pd.DataFrame) -> Dict[str, Dict]:
    """
    Look for plausible location entries in a dataframe.
    Flexible column matching for name/lat/lon.
    Returns dict keyed by normalized name.
    """
    results = {}
    if df is None or df.empty:
        return results

    df = df.copy()
    df.columns = [str(c).strip().lower() for c in df.columns]

    name_col = next((c for c in df.columns if c in ('name', 'place', 'location', 'building', 'department')), None)
    lat_col = next((c for c in df.columns if c in ('latitude', 'lat', 'y')), None)
    lon_col = next((c for c in df.columns if c in ('longitude', 'lon', 'x')), None)
    alias_col = next((c for c in df.columns if c in ('aliases', 'alias', 'nicknames')), None)
    desc_col = next((c for c in df.columns if c in ('description', 'desc', 'details')), None)

    if not (name_col and lat_col and lon_col):
        return results

    for _, row in df.iterrows():
        try:
            name = str(row[name_col]).strip()
            lat = float(row[lat_col])
            lon = float(row[lon_col])
            raw_aliases = str(row[alias_col]) if alias_col and pd.notna(row.get(alias_col)) else ""
            desc = str(row[desc_col]) if desc_col and pd.notna(row.get(desc_col)) else ""
            key = name.lower()
            results[key] = {
                "name": name,
                "original_name": name,
                "lat": lat,
                "lon": lon,
                "desc": desc,
                "aliases": _normalize_aliases(raw_aliases, name)
            }
        except Exception:
            continue
    return results

# ------------------- Unstructured text location extraction -------------------
def extract_locations_from_text(text: str) -> Dict[str, Dict]:
    """
    Hunt for coordinate patterns in free text and return location entries.
    Patterns handled:
      - "Place - Lat: xx.x, Lon: yy.y"
      - "Place (xx.x, yy.y)"
      - "Place Latitude: xx.x Longitude: yy.y"
    """
    locations = {}
    if not text or len(text) < 40:
        return locations

    patterns = [
        r'([\w\s\&\-]{3,60}?)\s*-\s*Lat[:\s]*([-+]?\d{1,3}\.?\d+),?\s*Lon[:\s]*([-+]?\d{1,3}\.?\d+)',
        r'([\w\s\&\-]{3,60}?)\s*Latitude[:\s]*([-+]?\d{1,3}\.?\d+)[,;\s]*Longitude[:\s]*([-+]?\d{1,3}\.?\d+)',
        r'([\w\s\&\-]{3,60}?)\s*\(\s*([-+]?\d{1,3}\.?\d+)[,\s]+([-+]?\d{1,3}\.?\d+)\s*\)'
    ]

    for pat in patterns:
        for m in re.finditer(pat, text, flags=re.IGNORECASE):
            try:
                name, lat, lon = m.groups()
                name = name.strip()
                key = name.lower()
                if key not in locations:
                    locations[key] = {
                        "name": name.title(),
                        "original_name": name,
                        "lat": float(lat),
                        "lon": float(lon),
                        "desc": f"Extracted from document text.",
                        "aliases": _normalize_aliases("", name)
                    }
            except Exception:
                continue
    return locations

# ------------------- Sentence extraction (chunking) -------------------
def extract_sentences_from_documents(file_data: List[Dict[str, str]]) -> List[Dict[str, str]]:
    """
    Convert raw document texts into sentence-chunks for FAISS.
    Returns list of {'sentence':..., 'source':...}
    """
    chunks = []
    for item in file_data:
        text = item.get("text", "")
        src = item.get("source", "unknown")
        if not text or len(text) < MIN_SENTENCE_LENGTH:
            continue

        # Normalize whitespace & remove bracketed refs
        t = re.sub(r'\s+', ' ', text)
        t = re.sub(r'\[.*?\]', '', t)

        # Split sentences using NLTK if ready; fallback to regex
        try:
            if NLTK_READY:
                sents = nltk.sent_tokenize(t)
            else:
                sents = re.split(r'(?<!\w\.\w.)(?<![A-Z][a-z]\.)(?<=\.|\?)\s', t)
        except Exception:
            sents = re.split(r'(?<=\.|\?)\s', t)

        for s in sents:
            s_clean = s.strip()
            if len(s_clean) >= MIN_SENTENCE_LENGTH:
                # collapse multiple spaces/newlines
                s_clean = re.sub(r'\s+', ' ', s_clean)
                chunks.append({"sentence": s_clean, "source": src})

    return chunks

# ------------------- Index building & persistence -------------------
def build_and_save_index(corpus: List[Dict[str, str]], location_map: Dict[str, Dict[str, Any]]) -> Tuple[bool,int,int]:
    """
    Build embeddings for corpus and save FAISS index + corpus pickle + locations pickle.
    Returns (success, num_sentences, num_locations)
    """
    try:
        # Sanity
        if not corpus:
            # remove any existing index/corpus to reflect empty state
            if FAISS_INDEX_PATH.exists(): FAISS_INDEX_PATH.unlink()
            if CORPUS_PATH.exists(): CORPUS_PATH.unlink()
            _log_debug("No corpus provided; removed existing index/corpus if any.")
            return True, 0, len(location_map or {})

        if EMBED_MODEL is None:
            st.error("Embedding model is not loaded. Cannot build FAISS index.")
            return False, 0, len(location_map or {})

        # de-duplicate sentences
        unique_map = {}
        for item in corpus:
            txt = item.get("sentence", "").strip()
            if txt:
                unique_map[txt] = item
        unique_sentences = list(unique_map.values())
        texts = [u['sentence'] for u in unique_sentences]

        # compute embeddings
        embeddings = EMBED_MODEL.encode(texts, show_progress_bar=True)
        emb_arr = np.array(embeddings, dtype='float32')
        dim = emb_arr.shape[1]

        index = faiss.IndexFlatL2(dim)
        index.add(emb_arr)
        faiss.write_index(index, str(FAISS_INDEX_PATH))

        # persist corpus
        with open(CORPUS_PATH, "wb") as f:
            pickle.dump(unique_sentences, f)

        # persist locations
        if location_map:
            with open(LOCATION_DATA_PATH, "wb") as f:
                pickle.dump(location_map, f)

        return True, len(unique_sentences), len(location_map or {})
    except Exception as e:
        st.error(f"❌ Error while building/saving index: {e}")
        return False, 0, len(location_map or {})

def load_system_index_and_corpus() -> Tuple[Optional[faiss.Index], List[Dict[str,str]], Dict[str, Dict[str,Any]]]:
    """
    Load FAISS index, corpus pickle, and locations pickle if available.
    Returns (index_or_None, corpus_list, location_map)
    """
    idx = None
    corpus = []
    locs = {}

    try:
        if FAISS_INDEX_PATH.exists() and CORPUS_PATH.exists():
            idx = faiss.read_index(str(FAISS_INDEX_PATH))
            with open(CORPUS_PATH, "rb") as f:
                corpus = pickle.load(f)
            _log_debug(f"Loaded index with {len(corpus)} items")
    except Exception as e:
        st.error(f"⚠️ Failed to load FAISS index or corpus: {e}")
        idx = None
        corpus = []

    try:
        if LOCATION_DATA_PATH.exists():
            with open(LOCATION_DATA_PATH, "rb") as f:
                locs = pickle.load(f)
    except Exception as e:
        _log_debug(f"Failed to load locations pickle: {e}")
        locs = {}

    return idx, corpus, locs

_log_debug("Batch 3 loaded: file processing, extraction, chunking, index helpers")


# ------------------- Chat session persistence -------------------
def _session_filename(session_id: str) -> Path:
    safe = hashlib.sha1(session_id.encode("utf-8")).hexdigest()
    return CHAT_HISTORY_DIR / f"session_{safe}.json"

def save_chat_session(session_id: str, chat_history: List[Dict[str, Any]]) -> None:
    try:
        fname = _session_filename(session_id)
        data = {
            "session_id": session_id,
            "timestamp": datetime.datetime.utcnow().isoformat() + "Z",
            "chat_history": chat_history[-MAX_CHAT_HISTORY_ITEMS:]
        }
        with open(fname, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        _log_debug(f"Saved session {session_id} ({len(chat_history)} messages)")
    except Exception as e:
        _log_debug(f"Failed to save session {session_id}: {e}")

def load_chat_session(session_id: str) -> List[Dict[str, Any]]:
    try:
        fname = _session_filename(session_id)
        if not fname.exists():
            return []
        with open(fname, "r", encoding="utf-8") as f:
            data = json.load(f)
            return data.get("chat_history", [])
    except Exception as e:
        _log_debug(f"Failed to load session {session_id}: {e}")
        return []

# ------------------- Intent classification (strict) -------------------
# A conservative set of campus-related keywords. We will treat queries as campus-related
# only if (a) they contain these keywords, OR (b) retrieval yields highly similar corpus hits.
CAMPUS_KEYWORDS = {
    "campus","library","library hours","department","faculty","professor","prof","hod",
    "office","building","canteen","mess","hostel","admissions","exam","syllabus","course",
    "curriculum","scholarship","tution","tuition","fee","fees","timetable","timetable","lab",
    "laboratory","zoology","physics","chemistry","computer","it","information","technology",
    "sports","gym","auditorium","bus","transport","office hours","contact","phone","email",
    "department of","where is","location","directions","map","how to reach","address"
}

LOCATION_KEYWORDS = {"where","where is","location","map","navigate","directions","how to reach","near"}

def is_location_intent(query: str) -> bool:
    if not query: return False
    q = query.lower()
    # crude detection: look for question words + location keywords
    for phrase in LOCATION_KEYWORDS:
        if phrase in q:
            return True
    # also if the query explicitly asks "show map" or contains coordinate-like pattern
    if re.search(r'\b(lat|lon|latitude|longitude|coord|coordinate)\b', q):
        return True
    return False

def is_campus_related_strict(query: str, corpus: List[Dict[str,str]], index: Optional[faiss.Index]) -> bool:
    """
    Strict policy:
      - If query contains an explicit campus keyword -> True
      - Else, if retrieval against corpus yields highly similar chunks (score threshold) -> True
      - Otherwise -> False
    This helps block off-topic requests.
    """
    if not query:
        return False

    q = query.lower()
    # quick keyword check
    for kw in CAMPUS_KEYWORDS:
        if kw in q:
            return True

    # If we have a corpus and index, do a conservative semantic check:
    try:
        if EMBED_MODEL and index and corpus:
            q_emb = EMBED_MODEL.encode([query])
            q_arr = np.array(q_emb, dtype='float32')
            # search top 3
            D, I = index.search(q_arr, 3)
            # faiss returns squared L2 distances in D; smaller is better
            # Convert to a rough similarity score: sim = 1 / (1 + dist)
            # We'll accept only very close matches: dist < threshold
            for dist in D[0]:
                if math.isfinite(dist) and dist >= 0:
                    # Accept if L2 distance less than a conservatively low threshold
                    if dist < 0.5:  # tuned conservatively; depends on embedding scale
                        return True
    except Exception as e:
        _log_debug(f"Semantic campus-check failed: {e}")

    # default: not campus-related
    return False

# ------------------- Retrieval & re-ranking -------------------
def retrieve_chunks(query: str, corpus: List[Dict[str,str]], index: Optional[faiss.Index], top_k: int = MAX_RETRIEVE) -> List[Dict[str,str]]:
    """
    Retrieves relevant sentence chunks from corpus using FAISS and re-ranks them by cosine similarity
    using the same embed model. Returns up to top_k chunks as a list.
    """
    if not query or not corpus or index is None or EMBED_MODEL is None:
        return []

    try:
        q_emb = EMBED_MODEL.encode([query])
        q_arr = np.array(q_emb, dtype='float32')

        # Faiss search (L2 index) - returns distances and indices
        D, I = index.search(q_arr, top_k)
        indices = [int(i) for i in I[0] if i < len(corpus)]
        retrieved = [corpus[i] for i in indices]

        # If we got no items, return empty
        if not retrieved:
            return []

        # Re-rank using cosine similarity with embeddings (more robust)
        try:
            # Compute embeddings for retrieved texts
            texts = [r['sentence'] for r in retrieved]
            emb_retrieved = EMBED_MODEL.encode(texts)
            # Normalize vectors
            q_vec = np.array(q_emb, dtype='float32')[0]
            q_norm = q_vec / (np.linalg.norm(q_vec) + 1e-12)
            emb_arr = np.array(emb_retrieved, dtype='float32')
            norms = np.linalg.norm(emb_arr, axis=1, keepdims=True) + 1e-12
            emb_norm = emb_arr / norms
            # Cosine similarities
            sims = (emb_norm @ q_norm).tolist()
            # attach scores and sort
            scored = list(zip(sims, retrieved))
            scored.sort(key=lambda x: x[0], reverse=True)
            top = [item for score, item in scored[:top_k]]
            return top
        except Exception as e:
            _log_debug(f"Re-ranking failed, returning Faiss order: {e}")
            return retrieved
    except Exception as e:
        _log_debug(f"Retrieval failed: {e}")
        return []

# ------------------- LLM prompt template & caller -------------------
def _build_system_prompt(language: str, geo_context: str, distance_info: str) -> str:
    """
    Build a strict system prompt instructing the LLM to only answer from provided context.
    """
    location_section = f"LOCATION CONTEXT:\n{geo_context}\n" if geo_context else "No location context provided.\n"
    distance_section = f"\n{distance_info}\n" if distance_info else ""
    prompt = f"""
You are CampusGPT, a concise and factual campus assistant. Follow these rules strictly:
1. Answer only using the provided RELEVANT CAMPUS INFORMATION and LOCATION CONTEXT. Do not invent facts.
2. If the answer cannot be found in the provided context, reply exactly: "I do not have information on that."
3. Keep answers concise (1-4 short paragraphs). Use bullet points for lists (faculty, courses, contacts).
4. If a user asks "where" or similar and a location is available, provide a brief natural answer followed by coordinates in the format (Lat: XX.XXXX, Lon: YY.YYYY).
5. Use the detected language for the reply. If language is unknown, default to English.
6. Do not mention "source", "document", "uploaded file", or "according to the documents".

LANGUAGE: {language}

{location_section}
{distance_section}
"""
    return prompt

def ask_llm_with_context(query: str, context_chunks: List[Dict[str,str]], geo_context: str, distance_info: str, groq_client: Any = GROQ_CLIENT, model_name: str = GROQ_MODEL_NAME, temperature: float = 0.2, max_tokens: int = 400) -> str:
    """
    Compose a strict prompt with retrieved context and call Groq LLM client.
    If Groq client not configured, return an informative message.
    """
    if groq_client is None:
        return "The AI assistant is currently offline. Please contact admin to configure the LLM."

    # Detect language
    try:
        lang_code = detect(query)
        lang_map = {'en':'English','ur':'Urdu','hi':'Hindi','es':'Spanish','fr':'French'}
        language = lang_map.get(lang_code, 'English')
    except LangDetectException:
        language = "English"

    system_prompt = _build_system_prompt(language, geo_context, distance_info)
    # Join context chunks into a coherent context block
    context_text = "\n".join([f"- {c['sentence']}" for c in context_chunks]) if context_chunks else ""

    full_prompt = f"""
{system_prompt}

---
RELEVANT CAMPUS INFORMATION:
{context_text if context_text else 'No relevant campus information was found.'}
---
USER QUESTION:
{query}

Provide a concise, factual response in {language}.
"""

    try:
        # Groq client wrapper
        resp = groq_client.chat.completions.create(
            model=model_name,
            messages=[{"role":"user","content": full_prompt}],
            temperature=temperature,
            max_tokens=max_tokens
        )
        # Navigate response structure safely
        content = ""
        try:
            content = resp.choices[0].message.content
        except Exception:
            # Some clients may return slightly different shapes
            content = getattr(resp, "text", str(resp))
        return content.strip()
    except Exception as e:
        _log_debug(f"LLM call failed: {e}")
        return "I apologize — I encountered an error while generating a response. Please try again later."

_log_debug("Batch 4 loaded: Intent classification, retrieval, re-ranking, LLM prompt logic")




# Map tile options (label -> folium tile name)
MAP_TILES = {
    "Satellite (Esri)": "Esri.WorldImagery",
    "Carto (Light)": "CartoDB positron",
    "OpenStreetMap": "OpenStreetMap",
    "Stamen Terrain": "Stamen Terrain"
}

# Default map display settings
DEFAULT_MAP_TILE = "Satellite (Esri)"
DEFAULT_MAP_ZOOM = 18
MAP_WIDTH = 900
MAP_HEIGHT = 480

# session_state keys used:
# - "map_tile" (str): current tile selection
# - "last_map_location" (dict): last focused location dict {name, lat, lon, desc}
# - "show_map" (bool): whether map should be displayed (set when location intent detected)

if "map_tile" not in st.session_state:
    st.session_state.map_tile = DEFAULT_MAP_TILE
if "last_map_location" not in st.session_state:
    st.session_state.last_map_location = None
if "show_map" not in st.session_state:
    st.session_state.show_map = False

def _build_marker_popup_html(loc: Dict[str, Any]) -> str:
    """
    Build HTML for marker popup containing description, a focus button (handled client-side as map control),
    and a Google Maps navigation link.
    """
    lat = loc['lat']
    lon = loc['lon']
    name = loc.get('original_name', loc.get('name', 'Location'))
    desc = loc.get('desc', '')
    maps_url = f"https://www.google.com/maps/dir/?api=1&destination={lat},{lon}"
    html = f"""
    <div style="width: 260px; font-family: Arial, sans-serif;">
      <h4 style="margin:0 0 6px 0;">{name}</h4>
      <div style="font-size:12px; color:#333; margin-bottom:8px;">{desc}</div>
      <div style="display:flex; gap:6px;">
        <a href="{maps_url}" target="_blank"
           style="padding:6px 10px; background:#1976d2; color:#fff; text-decoration:none; border-radius:4px; font-size:13px;">
          Navigate (Google Maps)
        </a>
      </div>
    </div>
    """
    return html

def create_expanded_map(locations: List[Dict[str, Any]], tile_label: Optional[str] = None, focus_idx: int = 0) -> Optional[folium.Map]:
    """
    Create a folium map centered on the mean coordinates of provided locations.
    If multiple locations exist, center on their mean; else center on single location.
    tile_label must be a key from MAP_TILES.
    focus_idx indicates which location to put emphasis on (0-based).
    """
    if not locations:
        return None

    tile_choice = tile_label or st.session_state.get("map_tile", DEFAULT_MAP_TILE)
    tile_name = MAP_TILES.get(tile_choice, list(MAP_TILES.values())[0])

    lats = [loc['lat'] for loc in locations]
    lons = [loc['lon'] for loc in locations]
    center = [float(np.mean(lats)), float(np.mean(lons))]

    # Choose zoom: default high zoom for single location, slightly wider for multiple
    zoom = DEFAULT_MAP_ZOOM if len(locations) == 1 else max(14, DEFAULT_MAP_ZOOM - 2)

    m = folium.Map(location=center, zoom_start=zoom, tiles=tile_name)

    # Add a marker for each location
    for i, loc in enumerate(locations):
        popup_html = _build_marker_popup_html(loc)
        folium.Marker(
            location=[loc['lat'], loc['lon']],
            popup=folium.Popup(popup_html, max_width=300),
            tooltip=loc.get('original_name', loc.get('name', '')).title()
        ).add_to(m)

    # Add simple locate control (zoom to user's location) - optional fallback
    try:
        from folium.plugins import LocateControl
        LocateControl(auto_start=False).add_to(m)
    except Exception:
        pass

    return m

def show_location_ui(loc: Dict[str, Any], context_text: Optional[str] = None):
    """
    Renders the map and location UI in the chat flow. Keeps the map expanded and provides controls:
    - Tile mode selector
    - Focus button (re-centers map on this marker)
    - Persistent display in session
    """
    st.session_state.show_map = True
    st.session_state.last_map_location = loc

    col_map_controls, col_map = st.columns([1, 3], gap="small")

    with col_map_controls:
        st.markdown("**Map options**")
        tile = st.selectbox("Map Mode", options=list(MAP_TILES.keys()), index=list(MAP_TILES.keys()).index(st.session_state.map_tile) if st.session_state.map_tile in MAP_TILES else 0, key="map_mode_select")
        # sync session value
        st.session_state.map_tile = tile

        if st.button("Focus on location"):
            # clicking this will simply re-render map centered on single loc (handled below by using last_map_location)
            st.session_state.last_map_location = loc
            st.experimental_rerun()

        st.markdown(f"**Coordinates:** `{loc['lat']:.6f}, {loc['lon']:.6f}`")
        if loc.get("desc"):
            st.markdown(f"**Details:** {loc.get('desc')}")

    with col_map:
        # Always expanded map for location queries
        map_obj = create_expanded_map([loc], tile_label=st.session_state.map_tile, focus_idx=0)
        if map_obj:
            st_folium(map_obj, width=MAP_WIDTH, height=MAP_HEIGHT)

# Helper to show the last map (if any) in a persistent area (e.g., when re-rendering chat history)
def show_last_map_if_any():
    if st.session_state.get("show_map") and st.session_state.get("last_map_location"):
        loc = st.session_state.last_map_location
        # Render a compact header + map
        st.markdown("---")
        st.markdown("### 📍 Last shown location")
        show_location_ui(loc)

# Utility: compute human-friendly distance if two locations present (kept from earlier plan)
def compute_distance_between_two(l1: Dict[str,Any], l2: Dict[str,Any]) -> str:
    try:
        c1 = (l1['lat'], l1['lon'])
        c2 = (l2['lat'], l2['lon'])
        d = geodesic(c1, c2)
        if d.kilometers >= 1:
            return f"Approximately {d.kilometers:.1f} km apart."
        else:
            return f"Approximately {d.meters:.0f} meters apart."
    except Exception:
        return ""

_log_debug("Batch 5 loaded: Map UI + folium helpers")



# ------------ Helper UI bits ------------
def _get_admin_password_input(key="admin_pass"):
    return st.sidebar.text_input("Admin password", type="password", key=key)

def _format_bot_message(text: str):
    # simple wrapper to make bot messages consistent (can extend with markdown styling)
    return text

def _strict_offtopic_reply(language: str = "English") -> str:
    # Keep it short and strict (translated if desired)
    # For now, English only. Could localize based on language detection.
    return "I can only answer campus-related questions. Please ask about campus locations, facilities, departments, services, or policies."

# ------------ Admin page ------------
def admin_page():
    st.title("🔧 Admin Portal")
    # Authentication
    if not st.session_state.get("admin_authenticated", False):
        st.write("Admin login required")
        pwd = st.text_input("Enter admin password", type="password", key="admin_login")
        if st.button("Login"):
            if pwd == ADMIN_PASSWORD:
                st.session_state.admin_authenticated = True
                st.success("Admin authenticated")
                st.experimental_rerun()
            else:
                st.error("Incorrect password")
        return

    st.success("✅ Admin access granted")
    tabs = st.tabs(["Upload & Process", "Locations CSV", "System Status", "Chat Histories", "Reset Index"])

    # --- Upload & Process Tab ---
    with tabs[0]:
        st.header("Upload & Process Documents")
        uploaded = st.file_uploader("Upload PDF, TXT, CSV, XLSX files (multiple allowed)", type=['pdf','txt','csv','xlsx','xls'], accept_multiple_files=True)
        if uploaded:
            st.info(f"{len(uploaded)} file(s) ready to process.")
        if st.button("Process Uploaded Files"):
            if not uploaded:
                st.warning("Please select files to upload first.")
            else:
                with st.spinner("Processing files and building index..."):
                    file_data, locs_from_files = process_uploaded_files(uploaded)
                    # Extract free-text location patterns from combined text
                    combined_text = " ".join([d['text'] for d in file_data])
                    text_locs = extract_locations_from_text(combined_text)
                    # Merge location maps (file-structured -> text-found)
                    merged_loc_map = {**LOCATION_MAP, **locs_from_files, **text_locs}
                    # Build sentence chunks
                    chunks = extract_sentences_from_documents(file_data)
                    success, n_sentences, n_locs = build_and_save_index(chunks, merged_loc_map)
                    # reload global location map cache
                    try:
                        # Update LOCATION_MAP persisted file and in-memory ref (simple approach)
                        with open(LOCATION_DATA_PATH, "wb") as f:
                            pickle.dump(merged_loc_map, f)
                        # refresh module-level cache (clear cached resource by re-calling get_location_map)
                        st.session_state.location_map = merged_loc_map
                        st.success(f"Processed: {n_sentences} knowledge items and {n_locs} locations.")
                    except Exception as e:
                        st.error(f"Error saving merged locations: {e}")

    # --- Locations CSV Tab ---
    with tabs[1]:
        st.header("Upload / Edit locations CSV")
        st.info("CSV should contain at least: name, latitude, longitude. Optional: aliases, description")
        uploaded_csv = st.file_uploader("Upload locations CSV (overwrites current)", type=['csv'], key="locations_csv_upload")
        if uploaded_csv:
            if st.button("Save Locations CSV"):
                # Save to LOCATIONS_DIR
                save_path = LOCATIONS_DIR / "locations.csv"
                with open(save_path, "wb") as f:
                    f.write(uploaded_csv.getbuffer())
                # Reload and persist processed pickle
                new_map = load_locations_from_csv(save_path)
                # update persisted pickle
                with open(LOCATION_DATA_PATH, "wb") as f:
                    pickle.dump(new_map, f)
                # update runtime map
                st.session_state.location_map = new_map
                st.success(f"Saved and loaded {len(new_map)} locations.")
        # Allow quick preview
        if (LOCATIONS_DIR / "locations.csv").exists():
            if st.checkbox("Preview current locations CSV"):
                try:
                    df_preview = pd.read_csv(LOCATIONS_DIR / "locations.csv")
                    st.dataframe(df_preview)
                except Exception as e:
                    st.error(f"Failed to preview CSV: {e}")

    # --- System Status Tab ---
    with tabs[2]:
        st.header("System Status")
        idx, corpus, loc_map = load_system_index_and_corpus()
        st.metric("Indexed knowledge items", len(corpus))
        st.metric("Indexed locations", len(loc_map))
        st.markdown(f"- FAISS index present: {'Yes' if FAISS_INDEX_PATH.exists() else 'No'}")
        st.markdown(f"- Corpus file present: {'Yes' if CORPUS_PATH.exists() else 'No'}")
        st.markdown(f"- Locations persisted: {'Yes' if LOCATION_DATA_PATH.exists() else 'No'}")
        # show last updated times
        def _file_mtime(path: Path):
            try:
                return datetime.datetime.utcfromtimestamp(path.stat().st_mtime).isoformat() + "Z"
            except Exception:
                return "N/A"
        st.markdown(f"- FAISS mtime: {_file_mtime(FAISS_INDEX_PATH)}")
        st.markdown(f"- Corpus mtime: {_file_mtime(CORPUS_PATH)}")
        st.markdown(f"- Locations mtime: {_file_mtime(LOCATION_DATA_PATH)}")

    # --- Chat Histories Tab ---
    with tabs[3]:
        st.header("Chat Histories")
        files = sorted(glob.glob(str(CHAT_HISTORY_DIR / "session_*.json")), reverse=True)
        if not files:
            st.info("No chat sessions found.")
        else:
            session_choice = st.selectbox("Select session file", options=files, format_func=lambda p: Path(p).name)
            if st.button("Load session"):
                try:
                    with open(session_choice, "r", encoding="utf-8") as f:
                        data = json.load(f)
                    chat = data.get("chat_history", [])
                    st.markdown(f"**Session ID:** {data.get('session_id')}")
                    st.markdown(f"**Saved at (UTC):** {data.get('timestamp')}")
                    st.markdown("----")
                    for m in chat:
                        role = m.get("role", "user")
                        content = m.get("content", "")
                        if role == "user":
                            st.markdown(f"**User:** {content}")
                        else:
                            st.markdown(f"**Assistant:** {content}")
                    if st.button("Download session as JSON"):
                        st.download_button("Download JSON", data=json.dumps(data, ensure_ascii=False, indent=2), file_name=Path(session_choice).name)
                except Exception as e:
                    st.error(f"Failed to load session: {e}")

    # --- Reset Index Tab ---
    with tabs[4]:
        st.header("Reset Index & Data")
        st.warning("This will remove FAISS index, corpus pickle, and persisted locations. This cannot be undone.")
        if st.button("Reset Index and Data"):
            confirm = st.checkbox("I confirm resetting the index and deleting persisted data")
            if confirm:
                # remove files
                for p in (FAISS_INDEX_PATH, CORPUS_PATH, LOCATION_DATA_PATH):
                    try:
                        if p.exists(): p.unlink()
                    except Exception as e:
                        st.error(f"Failed removing {p}: {e}")
                # optionally clear documents
                if st.checkbox("Also delete uploaded documents in data/documents"):
                    try:
                        for f in DOCUMENTS_DIR.iterdir():
                            if f.is_file(): f.unlink()
                        st.success("Deleted uploaded documents.")
                    except Exception as e:
                        st.error(f"Failed to delete documents: {e}")
                # clear session-level caches
                st.session_state.pop("location_map", None)
                st.success("Index and persisted data reset. Please upload new documents and locations.")

# ------------ User page (chat) ------------
def user_page():
    # Minimal top bar with small admin link
    st.markdown(
        """
        <div style="display:flex; align-items:center; justify-content:space-between;">
            <div style="display:flex; align-items:center; gap:12px;">
                <h2 style="margin:0;">CampusGPT — Your Campus Assistant</h2>
                <span style="color:gray;">Ask about campus locations, services, departments, and policies.</span>
            </div>
            <div>
                <a href="#" onclick="document.getElementById('admin-login-anchor').click();" style="font-size:14px;">Admin Login</a>
                <button id="admin-login-anchor" style="display:none;"></button>
            </div>
        </div>
        """, unsafe_allow_html=True
    )

    # Hide sidebar for users to keep UI clean
    st.query_params()  # noop to stabilize ui layout; sidebar hidden by default as configured earlier

    # Load system artifacts
    index, corpus, persisted_locs = load_system_index_and_corpus()
    # prefer session location_map if admin just uploaded
    runtime_loc_map = st.session_state.get("location_map", None) or (persisted_locs or LOCATION_MAP or {})
    # if everything empty -> show no data message
    system_ready = (index is not None and corpus) or bool(runtime_loc_map)

    if not system_ready:
        st.markdown("### ⚠️ No campus data available")
        st.info("No campus data available. Please contact admin.")
        return

    # Setup session id & history
    if "session_id" not in st.session_state:
        st.session_state.session_id = str(uuid.uuid4())
    if "chat_history" not in st.session_state:
        st.session_state.chat_history = load_chat_session(st.session_state.session_id) or []
    # trim if large
    if len(st.session_state.chat_history) > MAX_CHAT_HISTORY_ITEMS:
        st.session_state.chat_history = st.session_state.chat_history[-MAX_CHAT_HISTORY_ITEMS:]

    # Display previous messages in chat-like style
    chat_container = st.container()
    with chat_container:
        for m in st.session_state.chat_history:
            role = m.get("role", "user")
            content = m.get("content", "")
            if role == "user":
                st.chat_message("user").write(content)
            else:
                st.chat_message("assistant").write(content)

    # Show last map if available (persistent)
    show_last_map_if_any()

    # Chat input
    prompt = st.chat_input("Ask about the campus...")
    if prompt:
        # Append user message to session
        st.session_state.chat_history.append({"role":"user","content":prompt})
        save_chat_session(st.session_state.session_id, st.session_state.chat_history)

        # Immediately show user message in UI
        st.chat_message("user").write(prompt)

        # Determine language & intent
        try:
            lang_code = detect(prompt)
            lang_map = {'en':'English','ur':'Urdu','hi':'Hindi','es':'Spanish','fr':'French'}
            lang = lang_map.get(lang_code, 'English')
        except LangDetectException:
            lang = "English"

        # Strict domain check
        campus_ok = is_campus_related_strict(prompt, corpus, index)
        if not campus_ok:
            reply = "I can only answer campus-related questions. Please ask about campus locations, facilities, departments, services, or policies."
            st.chat_message("assistant").write(reply)
            st.session_state.chat_history.append({"role":"assistant","content":reply})
            save_chat_session(st.session_state.session_id, st.session_state.chat_history)
            return

        # Check if location intent
        if is_location_intent(prompt):
            # Try to match in locations CSV / map
            matched = fuzzy_match_location(prompt, runtime_loc_map, score_cutoff=75)
            if not matched:
                # If no match in CSV, still attempt to find via retrieval (some docs may contain coords)
                # Attempt retrieval then LLM answer without map
                chunks = retrieve_chunks(prompt, corpus, index, top_k=MAX_RETRIEVE)
                llm_answer = ask_llm_with_context(prompt, chunks, "", "", GROQ_CLIENT)
                st.chat_message("assistant").write(llm_answer)
                st.session_state.chat_history.append({"role":"assistant","content":llm_answer})
                save_chat_session(st.session_state.session_id, st.session_state.chat_history)
                return
            # If matched, prepare geo_context and distance_info
            geo_ctx = f"- **{matched.get('original_name', matched.get('name'))}**: (Lat: {matched['lat']:.6f}, Lon: {matched['lon']:.6f})"
            # get chunks (augment answer with doc info if available)
            chunks = retrieve_chunks(prompt, corpus, index, top_k=MAX_RETRIEVE)
            dist_info = ""
            # If user previously had another location in session, compute distance
            prev_loc = st.session_state.get("last_map_location")
            if prev_loc:
                try:
                    dist_info = compute_distance_between_two(prev_loc, matched)
                except Exception:
                    dist_info = ""
            # Ask LLM (will include instruction to include coords if available)
            answer = ask_llm_with_context(prompt, chunks, geo_ctx, dist_info, GROQ_CLIENT)
            # Natural reply first
            st.chat_message("assistant").write(answer)
            st.session_state.chat_history.append({"role":"assistant","content":answer})
            save_chat_session(st.session_state.session_id, st.session_state.chat_history)
            # Then show map expanded and controls (always)
            show_location_ui(matched, context_text=None)
            return

        # Non-location campus query: retrieve and ask LLM
        chunks = retrieve_chunks(prompt, corpus, index, top_k=MAX_RETRIEVE)
        answer = ask_llm_with_context(prompt, chunks, "", "", GROQ_CLIENT)
        st.chat_message("assistant").write(answer)
        st.session_state.chat_history.append({"role":"assistant","content":answer})
        save_chat_session(st.session_state.session_id, st.session_state.chat_history)
        return

# ------------ Main app ------------
def main():
    # Top-level layout: sidebar visible only for admin; user page is clean
    if "admin_authenticated" not in st.session_state:
        st.session_state.admin_authenticated = False

    # Show small admin login link in top-right via expander for safety
    with st.sidebar:
        st.markdown("## Navigation")
        mode = st.radio("Mode", options=["User Chat", "Admin Portal"], index=0)
        if st.button("Clear Chat History (local)"):
            if "chat_history" in st.session_state:
                st.session_state.chat_history = []
                save_chat_session(st.session_state.session_id, [])
                st.experimental_rerun()

    if mode == "Admin Portal":
        admin_page()
    else:
        user_page()

if __name__ == "__main__":
    main()


# ------------------- UI polishing (CSS & avatars) -------------------
CHAT_CSS = """
<style>
/* Page width */
main .block-container {
  max-width: 900px;
  margin: auto;
  padding-top: 10px;
}

/* Chat bubbles */
.streamlit-chat-user {
  display: flex;
  justify-content: flex-end;
  margin-bottom: 8px;
}
.streamlit-chat-bot {
  display: flex;
  justify-content: flex-start;
  margin-bottom: 8px;
}
.user-bubble {
  background: linear-gradient(90deg,#2979ff,#1e88e5);
  color: white;
  padding: 10px 14px;
  border-radius: 14px;
  max-width: 78%;
  box-shadow: 0 1px 2px rgba(0,0,0,0.08);
  font-size: 14px;
}
.bot-bubble {
  background: #f1f5f9;
  color: #0b1220;
  padding: 10px 14px;
  border-radius: 14px;
  max-width: 78%;
  font-size: 14px;
  box-shadow: 0 1px 2px rgba(0,0,0,0.04);
}
.small-muted {
  font-size:12px; color:#6b7280; margin-top:6px;
}
/* Make chat container scroll nicely */
.chat-area {
  max-height: 60vh;
  overflow: auto;
  padding: 6px;
}
/* Make the input feel sticky */
.stChatInput {
  position: sticky;
  bottom: 0;
  background: transparent;
}
</style>
"""

# inject CSS once
st.markdown(CHAT_CSS, unsafe_allow_html=True)

# small helper wrappers for styled chat message rendering
def render_user_bubble(text: str):
    safe = html.escape(text)
    st.markdown(f'<div class="streamlit-chat-user"><div class="user-bubble">{safe}</div></div>', unsafe_allow_html=True)

def render_bot_bubble(text: str):
    # allow minimal markdown and links
    safe = text.replace("\n", "<br>")
    st.markdown(f'<div class="streamlit-chat-bot"><div class="bot-bubble">{safe}</div></div>', unsafe_allow_html=True)

# ------------------- Structured-answer prompt builder -------------------
def build_structured_answer_prompt(entity_name: str, language: str = "English") -> str:
    """
    Prompt instructing LLM to extract structured information about an entity (department, facility).
    The calling code will append the retrieved context below this system instruction.
    """
    p = f"""
You are CampusGPT, a concise and factual campus assistant. The user asked about: "{entity_name}".

Task:
- Using ONLY the provided RELEVANT CAMPUS INFORMATION below, extract and format available details into the following sections when present:
  1. Short description (1-2 sentences)
  2. Head of Department (HOD) or in-charge (name & title)
  3. Some faculty/staff (list up to 5 names)
  4. Courses / Services offered (bullet list)
  5. Office hours / Timings
  6. Contact information (email, phone)
  7. Any other useful notes (e.g., prerequisites, eligibility, fees)

Formatting rules:
- Use Markdown.
- Use bold for section titles.
- If a section cannot be answered from the provided context, omit it (do NOT invent).
- Keep the whole answer under 400 words if possible.
- At the end, add a one-line prompt: "Would you like the location on the map?" (only this sentence)

LANGUAGE: {language}

Now below provide the RELEVANT CAMPUS INFORMATION.
"""
    return p

# Helper to decide if a query should use structured extraction (like "about X", "tell me about X")
def is_structured_query(query: str) -> bool:
    q = query.lower().strip()
    # naive heuristics: starts with "about", "tell me about", contains "info on", "information on"
    if q.startswith("about ") or q.startswith("tell me about ") or "information on" in q or q.endswith(" department") or q.endswith(" dept") or q.startswith("who is the head of"):
        return True
    # also short queries that are just a department or entity name (one or two words)
    if len(q.split()) <= 3 and any(word in q for word in ["department", "dept", "library", "canteen", "office", "centre", "center", "lab"]):
        return True
    return False

# ------------------- Admin utilities: export/import index & corpus -------------------
def admin_export_index_and_corpus():
    exports = {}
    if FAISS_INDEX_PATH.exists():
        try:
            with open(FAISS_INDEX_PATH, "rb") as f:
                exports['faiss_index'] = f.read()
        except Exception as e:
            st.error(f"Failed reading FAISS file: {e}")
    if CORPUS_PATH.exists():
        try:
            with open(CORPUS_PATH, "rb") as f:
                exports['corpus'] = f.read()
        except Exception as e:
            st.error(f"Failed reading corpus pickle: {e}")
    if LOCATION_DATA_PATH.exists():
        try:
            with open(LOCATION_DATA_PATH, "rb") as f:
                exports['locations'] = f.read()
        except Exception as e:
            st.error(f"Failed reading locations pickle: {e}")

    if not exports:
        st.info("Nothing to export (no index/corpus/locations present).")
        return

    # prepare a zip-like single binary blob via pickle (for simplicity)
    try:
        blob = pickle.dumps(exports)
        st.download_button("Download system export (.pkl)", data=blob, file_name="campusgpt_export.pkl")
    except Exception as e:
        st.error(f"Failed preparing export: {e}")

def admin_import_exported_blob(uploaded_blob):
    try:
        data = pickle.load(uploaded_blob)
        # write files back
        if 'faiss_index' in data:
            with open(FAISS_INDEX_PATH, "wb") as f:
                f.write(data['faiss_index'])
        if 'corpus' in data:
            with open(CORPUS_PATH, "wb") as f:
                f.write(data['corpus'])
        if 'locations' in data:
            with open(LOCATION_DATA_PATH, "wb") as f:
                f.write(data['locations'])
        st.success("Imported system data. You may want to reload the app to refresh in-memory caches.")
    except Exception as e:
        st.error(f"Import failed: {e}")

# ------------------- Small UX improvements -------------------
def _simulate_typing_delay(min_delay=0.35, max_delay=0.8):
    # small randomized delay to simulate typing (keeps snappy; do not make it long)
    try:
        _time.sleep(min_delay)
    except Exception:
        pass

def _safe_write_assistant(content: str):
    """
    Ensure assistant output is reasonably sized before rendering. If large, allow download and show truncation.
    """
    MAX_CHARS = 4000
    if len(content) <= MAX_CHARS:
        render_bot_bubble(content)
    else:
        # show first part and allow download of full response
        render_bot_bubble(content[:MAX_CHARS] + "\n\n*(Output truncated — download full response)*")
        st.download_button("Download full assistant response", data=content, file_name="assistant_response.txt")

# ------------------- Hook into admin UI (small) -------------------
# Add export/import controls to admin page by monkey-patching small helper (used in admin_page)
def _admin_extra_tools():
    st.subheader("Maintenance & Exports")
    if st.button("Export FAISS + Corpus + Locations"):
        admin_export_index_and_corpus()
    uploaded_blob = st.file_uploader("Import system export (.pkl) to restore", type=['pkl'])
    if uploaded_blob:
        if st.button("Import uploaded export"):
            admin_import_exported_blob(uploaded_blob)
